import base64
import logging
import time
from copy import copy, deepcopy
from io import BytesIO

from odoo import fields, models
from odoo.modules.module import get_module_resource

_logger = logging.getLogger(__name__)
try:
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from openpyxl.utils.exceptions import IllegalCharacterError
except ImportError:
    _logger.debug(
        'Cannot import "openpyxl". Please make sure it is installed.')


class SLocationInventoryReportFile(models.TransientModel):
    _name = 's.location.inventory.report.file'

    file_name = fields.Char()
    file = fields.Binary()


class SLocationInventoryReport(models.TransientModel):
    _name = 's.location.inventory.report'

    def default_location(self):
        valid_location = []
        locations = []
        user = self.env.user
        if user.has_group('advanced_sale.s_boo_group_administration') or user.has_group(
                'advanced_sale.s_boo_group_kiem_ke') or user.has_group('advanced_sale.s_boo_group_hang_hoa') or user.has_group('advanced_sale.s_boo_group_dieu_phoi'):
            locations = self.env['stock.location'].search([])
        elif user.has_group('advanced_sale.s_boo_group_thu_ngan'):
            locations = self.env['stock.location'].search(
                [('warehouse_id_store.employee_ids', 'in', user.employee_ids.ids)])
        for location in locations:
            valid_location.append(location.id)
        return [(6, 0, valid_location)]

    location_default_ids = fields.Many2many('stock.location', 'table_location_default', default=default_location)
    location_ids = fields.Many2many('stock.location', string='Nhà kho')

    def confirm_export(self):
        wb = load_workbook(get_module_resource('advanced_inventory', 'template',
                                               'bao_cao_ton_kho_dia_diem_template.xlsx'))
        ws = wb['Sheet1']
        self.fill_data(workbook=wb, worksheet=ws)
        content = BytesIO()
        wb.save(content)
        out = base64.b64encode(content.getvalue())
        view = self.env.ref('advanced_inventory.s_location_inventory_report_file_view_form')
        self.env['s.location.inventory.report.file'].sudo().search([]).unlink()
        file_output = self.env['s.location.inventory.report.file'].sudo().create({
            'file': out,
            'file_name': "Báo cáo tồn kho theo địa điểm.xlsx"
        })
        content.close()
        return {
            'name': "Báo cáo kho",
            'type': 'ir.actions.act_window',
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 's.location.inventory.report.file',
            'target': 'new',
            'view_id': view.id,
            'res_id': file_output.id
        }

    def fill_data(self, workbook, worksheet):
        start_time = time.time()
        products = self._cr.execute(
            'select product_product.id as id,'
            'product_template.name as name,'
            'product_product.default_code as default_code,'
            'product_template.ma_san_pham as ma_san_pham,'
            'product_product.ma_vat_tu as ma_vat_tu,'
            'product_template.ma_cu as ma_cu,'
            'product_template.list_price as list_price,'
            'product_product.kich_thuoc as kich_thuoc,'
            'ma_size from product_product join product_template on product_product.product_tmpl_id = product_template.id '
            'where product_product.barcode is not null')
        products = self._cr.dictfetchall()
        location_domain_list = [str(e) for e in self.location_ids.ids]
        location_domain_str = '(' + ','.join(location_domain_list) + ')'
        product_domain_list = [str(e['id']) for e in products]
        product_domain_str = '(' + ','.join(product_domain_list) + ')'
        all_stock_quant = self._cr.execute('select location_id,product_id,reserved_quantity,quantity from stock_quant where location_id in ' + location_domain_str + ' and product_id in ' + product_domain_str)
        all_stock_quant = self._cr.dictfetchall()
        all_stock_move_line = self._cr.execute("""select location_id,location_dest_id,product_id,product_uom_qty from stock_move_line where state like 'assigned' and location_dest_id in """ + location_domain_str + ' and product_id in ' + product_domain_str)
        all_stock_move_line = self._cr.dictfetchall()
        all_stock_scrap = self._cr.execute('select scrap_qty,product_id,location_id from stock_scrap where location_id in ' + location_domain_str + ' and product_id in ' + product_domain_str)
        all_stock_scrap = self._cr.dictfetchall()
        locations = self.location_ids
        title_style = copy(worksheet['M1']._style)
        # style = copy(worksheet['B2']._style)
        loc_col_index = 14
        location_dict = {}
        for location in locations:
            stock_warehouse = self.env['stock.warehouse'].sudo().search([('lot_stock_id', '=', location.id)])
            if stock_warehouse:
                worksheet.cell(row=1, column=loc_col_index)._style = title_style
                worksheet.cell(row=1,
                               column=loc_col_index).value = "Tồn có thể bán (" + location.name + ' ' + stock_warehouse.name + ")"
                loc_col_index += 1
                worksheet.cell(row=1, column=loc_col_index)._style = title_style
                worksheet.cell(row=1,
                               column=loc_col_index).value = "Tồn chờ xuất (" + location.name + ' ' + stock_warehouse.name + ")"
                loc_col_index += 1
                worksheet.cell(row=1, column=loc_col_index)._style = title_style
                worksheet.cell(row=1,
                               column=loc_col_index).value = "Tồn sắp về (" + location.name + ' ' + stock_warehouse.name + ")"
                loc_col_index += 1
                worksheet.cell(row=1, column=loc_col_index)._style = title_style
                worksheet.cell(row=1,
                               column=loc_col_index).value = "Tồn thực tế (" + location.name + ' ' + stock_warehouse.name + ")"
                loc_col_index += 1
                worksheet.cell(row=1, column=loc_col_index)._style = title_style
                worksheet.cell(row=1,
                               column=loc_col_index).value = "Phế liệu (" + location.name + ' ' + stock_warehouse.name + ")"
                loc_col_index += 1
            location_dict[location.id] = {
                'ton_co_the_ban': 0,
                'cho_xuat': 0,
                'ton_sap_ve': 0,
                'ton_thuc_te': 0,
                'phe_lieu': 0,
            }
        worksheet_data = {}
        for product in products:
            worksheet_data[product['id']] = {
                'default_code': product['default_code'],
                'ma_san_pham': product['ma_san_pham'],
                'ma_vat_tu': product['ma_vat_tu'],
                'ma_size': product['ma_size'],
                'ten_den_size': "ten den size",
                'ma_cu': product['ma_cu'],
                'ten_san_pham': product['name'],
                'size': product['kich_thuoc'],
                'gia_ban': product['list_price'],
                'tong_ton_co_the_ban': 0,
                'tong_cho_xuat': 0,
                'tong_ton_sap_ve': 0,
                'tong_ton_thuc_te': 0,
                'tong_phe_lieu': 0,
                'location': deepcopy(location_dict)
            }
        for stock_quant in all_stock_quant:
            worksheet_data[stock_quant['product_id']]['tong_ton_co_the_ban'] += stock_quant['quantity'] - stock_quant['reserved_quantity']
            worksheet_data[stock_quant['product_id']]['tong_cho_xuat'] += stock_quant['reserved_quantity']
            worksheet_data[stock_quant['product_id']]['tong_ton_thuc_te'] += stock_quant['quantity']
            worksheet_data[stock_quant['product_id']]['location'][stock_quant['location_id']]['ton_co_the_ban'] += stock_quant['quantity'] - stock_quant['reserved_quantity']
            worksheet_data[stock_quant['product_id']]['location'][stock_quant['location_id']]['cho_xuat'] += stock_quant['reserved_quantity']
            worksheet_data[stock_quant['product_id']]['location'][stock_quant['location_id']]['ton_thuc_te'] += stock_quant['quantity']
        for stock_scrap in all_stock_scrap:
            worksheet_data[stock_scrap['product_id']]['tong_phe_lieu'] += stock_scrap['scrap_qty']
            worksheet_data[stock_scrap['product_id']]['location'][stock_scrap['location_id']]['phe_lieu'] += stock_scrap['scrap_qty']
        for stock_move_line in all_stock_move_line:
            worksheet_data[stock_move_line['product_id']]['tong_ton_sap_ve'] += stock_move_line['product_uom_qty']
            # worksheet_data[stock_move_line['product_id']]['location'][stock_move_line['location_dest_id']]['ton_sap_ve'] += stock_move_line['qty_done']
            worksheet_data[stock_move_line['product_id']]['location'][stock_move_line['location_dest_id']]['ton_sap_ve'] += stock_move_line['product_uom_qty']
        if len(locations) > 0:
            if len(products) > 0:
                count = 0
                current_worksheet_row = 1
                for product in products:
                    count += 1
                    current_worksheet_row += 1
                    # if count > 10:
                    #     break
                    # print(str(count) + '/' + str(len(products)))
                    col = 14
                    for location in locations:
                        stock_warehouse_id = self.env['stock.warehouse'].sudo().search([('lot_stock_id', '=', location.id)])
                        if stock_warehouse_id:
                            # worksheet.cell(row=current_worksheet_row, column=col)._style = style
                            worksheet.cell(row=current_worksheet_row, column=col).value = worksheet_data[product['id']]['location'][location.id]['ton_co_the_ban']
                            col += 1
                            # worksheet.cell(row=current_worksheet_row, column=col)._style = style
                            worksheet.cell(row=current_worksheet_row, column=col).value = worksheet_data[product['id']]['location'][location.id]['cho_xuat']
                            col += 1
                            # worksheet.cell(row=current_worksheet_row, column=col)._style = style
                            worksheet.cell(row=current_worksheet_row, column=col).value = worksheet_data[product['id']]['location'][location.id]['ton_sap_ve']
                            col += 1
                            # worksheet.cell(row=current_worksheet_row, column=col)._style = style
                            worksheet.cell(row=current_worksheet_row, column=col).value = worksheet_data[product['id']]['location'][location.id]['ton_thuc_te']
                            col += 1
                            # worksheet.cell(row=current_worksheet_row, column=col)._style = style
                            worksheet.cell(row=current_worksheet_row, column=col).value = worksheet_data[product['id']]['location'][location.id]['phe_lieu']
                            col += 1
                    # worksheet.cell(row=current_worksheet_row, column=1)._style = style
                    worksheet.cell(row=current_worksheet_row, column=1).value = worksheet_data[product['id']]['default_code']
                    # worksheet.cell(row=current_worksheet_row, column=2)._style = style
                    worksheet.cell(row=current_worksheet_row, column=2).value = worksheet_data[product['id']]['ma_san_pham']
                    # worksheet.cell(row=current_worksheet_row, column=3)._style = style
                    worksheet.cell(row=current_worksheet_row, column=3).value = worksheet_data[product['id']]['ma_vat_tu']
                    # worksheet.cell(row=current_worksheet_row, column=4)._style = style
                    worksheet.cell(row=current_worksheet_row, column=4).value = worksheet_data[product['id']]['ma_size']
                    # worksheet.cell(row=current_worksheet_row, column=5)._style = style
                    worksheet.cell(row=current_worksheet_row, column=5).value = worksheet_data[product['id']]['ten_den_size']
                    # worksheet.cell(row=current_worksheet_row, column=6)._style = style
                    worksheet.cell(row=current_worksheet_row, column=6).value = worksheet_data[product['id']]['ma_cu']
                    # worksheet.cell(row=current_worksheet_row, column=7)._style = style
                    worksheet.cell(row=current_worksheet_row, column=7).value = worksheet_data[product['id']]['ten_san_pham']
                    # worksheet.cell(row=current_worksheet_row, column=8)._style = style
                    worksheet.cell(row=current_worksheet_row, column=8).value = worksheet_data[product['id']]['size']
                    # worksheet.cell(row=current_worksheet_row, column=9)._style = style
                    worksheet.cell(row=current_worksheet_row, column=9).value = worksheet_data[product['id']]['gia_ban']
                    # worksheet.cell(row=current_worksheet_row, column=10)._style = style
                    worksheet.cell(row=current_worksheet_row, column=10).value = worksheet_data[product['id']]['tong_ton_co_the_ban']
                    # worksheet.cell(row=current_worksheet_row, column=11)._style = style
                    worksheet.cell(row=current_worksheet_row, column=11).value = worksheet_data[product['id']]['tong_cho_xuat']
                    # worksheet.cell(row=current_worksheet_row, column=12)._style = style
                    worksheet.cell(row=current_worksheet_row, column=12).value = worksheet_data[product['id']]['tong_ton_sap_ve']
                    # worksheet.cell(row=current_worksheet_row, column=13)._style = style
                    worksheet.cell(row=current_worksheet_row, column=13).value = worksheet_data[product['id']]['tong_ton_thuc_te']
        print('total = ' + str(time.time() - start_time))

    def fill_data_b(self, workbook, worksheet):
        products = self.env['product.product'].sudo().search([('type', '=', 'product')])
        locations = self.location_ids
        title_style = copy(worksheet['M1']._style)
        style = copy(worksheet['B2']._style)
        loc_col_index = 14
        for location in locations:
            stock_warehouse = self.env['stock.warehouse'].sudo().search([('lot_stock_id', '=', location.id)])
            if stock_warehouse:
                worksheet.cell(row=1, column=loc_col_index)._style = title_style
                worksheet.cell(row=1,
                               column=loc_col_index).value = "Tồn có thể bán (" + location.name + ' ' + stock_warehouse.name + ")"
                loc_col_index += 1
                worksheet.cell(row=1, column=loc_col_index)._style = title_style
                worksheet.cell(row=1,
                               column=loc_col_index).value = "Tồn chờ xuất (" + location.name + ' ' + stock_warehouse.name + ")"
                loc_col_index += 1
                worksheet.cell(row=1, column=loc_col_index)._style = title_style
                worksheet.cell(row=1,
                               column=loc_col_index).value = "Tồn sắp về (" + location.name + ' ' + stock_warehouse.name + ")"
                loc_col_index += 1
                worksheet.cell(row=1, column=loc_col_index)._style = title_style
                worksheet.cell(row=1,
                               column=loc_col_index).value = "Tồn thực tế (" + location.name + ' ' + stock_warehouse.name + ")"
                loc_col_index += 1
        if len(locations) > 0:
            if len(products) > 0:
                for product in products:
                    worksheet.insert_rows(2)
                    tong_ton_sap_ve = 0
                    tong_cho_xuat = 0
                    tong_ton_co_the_ban = 0
                    tong_ton_thuc_te = 0
                    default_code = product.default_code if product.default_code else " "
                    ma_san_pham = product.ma_san_pham if product.ma_san_pham else " "
                    ma_vat_tu = product.ma_vat_tu if product.ma_vat_tu else " "
                    ten_den_size = "ten den size"
                    ma_cu = product.ma_cu if product.ma_cu else " "
                    ten_san_pham = product.name if product.name else " "
                    gia_ban = str(product.list_price) if product.list_price else " "
                    size = ""
                    ma_size = ""
                    if product.kich_thuoc:
                        size = product.kich_thuoc
                        product_attrs = product.product_template_attribute_value_ids
                        for product_size in product_attrs.attribute_id:
                            if product_size.type == "size":
                                for product_size_code in product_attrs:
                                    if product_size_code.product_attribute_value_id.code:
                                        ma_size = product_size_code.product_attribute_value_id.code
                    col = 14
                    for location in locations:
                        ton_sap_ve = 0
                        cho_xuat = 0
                        ton_co_the_ban = 0
                        ton_thuc_te = 0
                        data = self.env['stock.quant'].sudo().read_group([
                            ('product_id', '=', product.id),
                            ('location_id', '=', location.id)
                        ], ['location_id', 'quantity', 'reserved_quantity'], ['location_id'])
                        move_line_ids = self.env['stock.move.line'].sudo().search(
                            [('product_id', '=', product.id), ('location_dest_id', '=', location.id),
                             ('state', '=', 'assigned')])
                        if len(move_line_ids) > 0:
                            ton_sap_ve = sum([line.product_uom_qty for line in move_line_ids])
                        if len(data) > 0:
                            cho_xuat = data[0]['reserved_quantity']
                            ton_thuc_te = data[0]['quantity']
                            ton_co_the_ban = ton_thuc_te - cho_xuat
                        tong_ton_sap_ve += ton_sap_ve
                        tong_cho_xuat += cho_xuat
                        tong_ton_co_the_ban += ton_co_the_ban
                        tong_ton_thuc_te += ton_thuc_te
                        worksheet.cell(row=2, column=col)._style = style
                        worksheet.cell(row=2, column=col).value = ton_co_the_ban
                        col += 1
                        worksheet.cell(row=2, column=col)._style = style
                        worksheet.cell(row=2, column=col).value = cho_xuat
                        col += 1
                        worksheet.cell(row=2, column=col)._style = style
                        worksheet.cell(row=2, column=col).value = ton_sap_ve
                        col += 1
                        worksheet.cell(row=2, column=col)._style = style
                        worksheet.cell(row=2, column=col).value = ton_thuc_te
                        col += 1
                    worksheet.cell(row=2, column=1)._style = style
                    worksheet.cell(row=2, column=1).value = default_code
                    worksheet.cell(row=2, column=2)._style = style
                    worksheet.cell(row=2, column=2).value = ma_san_pham
                    worksheet.cell(row=2, column=3)._style = style
                    worksheet.cell(row=2, column=3).value = ma_vat_tu
                    worksheet.cell(row=2, column=4)._style = style
                    worksheet.cell(row=2, column=4).value = ma_size
                    worksheet.cell(row=2, column=5)._style = style
                    worksheet.cell(row=2, column=5).value = ten_den_size
                    worksheet.cell(row=2, column=6)._style = style
                    worksheet.cell(row=2, column=6).value = ma_cu
                    worksheet.cell(row=2, column=7)._style = style
                    worksheet.cell(row=2, column=7).value = ten_san_pham
                    worksheet.cell(row=2, column=8)._style = style
                    worksheet.cell(row=2, column=8).value = size
                    worksheet.cell(row=2, column=9)._style = style
                    worksheet.cell(row=2, column=9).value = gia_ban
                    worksheet.cell(row=2, column=10)._style = style
                    worksheet.cell(row=2, column=10).value = tong_ton_co_the_ban
                    worksheet.cell(row=2, column=11)._style = style
                    worksheet.cell(row=2, column=11).value = tong_cho_xuat
                    worksheet.cell(row=2, column=12)._style = style
                    worksheet.cell(row=2, column=12).value = tong_ton_sap_ve
                    worksheet.cell(row=2, column=13)._style = style
                    worksheet.cell(row=2, column=13).value = tong_ton_thuc_te
